

import openpyxl

def ReadData(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)    # file
    sheet=workbook[sheetname]            # sheet
    return sheet.cell(row=rownum,column=colnum).value


def WriteData(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)    # file
    sheet=workbook[sheetname]            # sheet
    sheet.cell(row=rownum,column=colnum).value = data   # enter data
    workbook.save(file)   # save file


def RowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)  # file
    sheet = workbook[sheetname]     # sheet
    return sheet.max_row

